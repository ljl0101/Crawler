from ..views import ticketsList

from scrapy.http.request import Request
import scrapy

import xlwt
import xlrd
from copy import copy
import openpyxl

import time

#! scrapy crawl ticketInfo


class stationsList(scrapy.Spider):
    name = 'ticketInfo'
    allowed_domains = ['www.kyfw.12306.cn']

    JSESSIONID = '7A134DDF98831A43EEE1E54415105F43'

    # & default
    param = {
        # 时间
        'leftTicketDTO.train_date': time.strftime("%Y-%m-%d", time.localtime(time.time())),
        'leftTicketDTO.from_station': 'WHN',  # 出发地
        'leftTicketDTO.to_station': 'BJP',  # 目的地
        'purpose_codes': 'ADULT'
    }

    # & default
    ticket_info = {
        "date": "2021-05-05",
        "start": "WHN",
        "arrive": "BJP",
        "order": "Z78",
        "start_time": "01:18",
        "arrive_time": "12:15",
        "pass_time": "10:57"
    }

    # & default
    prices = {
        'A9': 0,    # 商务座
        'M': 0,     # 一等座
        'WZ': 0,    # 二等座
        'A6': 0,    # 高级软卧
        'A4': 0,    # 软卧一等座
        '3': 0,     # 硬卧二等座
        '1': 0      # 硬座 / 无座
    }

    cookie = {
        'JSESSIONID': '7A134DDF98831A43EEE1E54415105F43',
        'tk': '7Q0g-f4mPZmlXTu0Kak5jdRHhu0wW37CM9ZLEgsdl1l0',
        'BIGipServerotn': '703595018.38945.0000',
        'BIGipServerpool_passport': '149160458.50215.0000',
        'RAIL_EXPIRATION': '1619537962659',
        'RAIL_DEVICEID': 'jiSrzl_X0nDcHgcH33Uxj6rqOZKeMUeLo-LvvXkEL2NI3Sg3CRTIWVf5WLR3B6Agcz2PFpot8k_F8XZYesRmrqhgFU6wU2SpmNXiRuX9ULJ9tqeg5WFQ7BsurDyw-GMap52GUNgC1b0VJtJacPUfjD0n0g1xmELp',
        'route': '495c805987d0f5c8c84b14f60212447d',
        '_jc_save_fromStation': '%u5317%u4EAC%2CBJP',
        '_jc_save_fromDate': '2021-04-24',
        '_jc_save_toDate': '2021-04-24',
        '_jc_save_wfdc_flag': 'dc',
        '_jc_save_toStation': '%u4E0A%u6D77%2CSHH',
        'uKey': '9f01dfaa51e7efc519ec6fa99f9c80aa0af85da3e7f70cec3aed53d10d7efac8'
    }

    ticket_url = 'https://kyfw.12306.cn/otn/leftTicket/queryTicketPrice'

    def __init__(self, name=None, **kwargs):
        super().__init__(name=name, **kwargs)
        tickets = ticketsList.ticketsList()
        tickets.main()

        self.param = tickets.param
        self.ticket_info = tickets.ticket_info
        status = tickets.status

        if status == 1:
            self.ticket_url += "?"
            # 提取param里的信息添加到ticket_url后面
            for key in self.param.keys():
                self.ticket_url += key + "=" + self.param[key] + "&"
            self.ticket_url = self.ticket_url[:-1]

    def start_requests(self):
        return [Request(
            self.ticket_url,
            callback=self.get_ticket_info,
            # & 获取火车站字典列表(resources/station.json)
            meta={
                'dont_redirect': True,
                'handle_httpstatus_list': [301, 302]
            },
            dont_filter=True,
            cookies=self.cookie,
            errback=self.error
        ),
        ]

        # https://kyfw.12306.cn/otn/leftTicket/query?leftTicketDTO.train_date=2021-04-21&leftTicketDTO.from_station=BJP&leftTicketDTO.to_station=TJP&purpose_codes=ADULT

    def error(self, response):
        print('fail')

    def get_ticket_info(self, response):
        try:
            response = response.json()
            for key in response['data'].keys():
                if(key in self.prices.keys()):
                    if(key[0] <= '9' and key[0] >= '0'):
                        self.prices[key] = eval(
                            str(response['data'][key])) / 10
                    else:
                        self.prices[key] = eval(str(response['data'][key][1:]))

            self.excelWrite(self.ticket_info, self.prices)

        except Exception:
            print('请勿直接关闭Python图形化界面')


    def excelWrite(self, ticket_info, prices):
        workbook = openpyxl.load_workbook("../resources/ticketInfo.xlsx")
        worksheet = workbook.worksheets[0]

        index = 1
        # 判断从第几行开始写入
        while worksheet.cell(index, 1).value != None:
            index += 1

        print(index)

        # 写入表格信息
        column = 1
        for info in ticket_info.values():
            worksheet.cell(index, column, info)
            column += 1

        for price in prices.values():
            worksheet.cell(index, column, price)
            column += 1

        worksheet.cell(index, column, price)

        workbook.save(filename="../resources/ticketInfo.xlsx")
